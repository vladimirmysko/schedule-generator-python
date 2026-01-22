"""Excel instructor schedule generator from JSON data."""

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .constants import get_slot_time_range

# Localization strings
INSTRUCTOR_STRINGS_KAZ = {
    "university": "Батыс Қазақстан инновациялық-технологиялық университеті",
    "schedule_title": "ОҚЫТУШЫ КЕСТЕСІ",
    "legend_both": "Екі апта",
    "legend_odd": "Тақ апта",
    "legend_even": "Жұп апта",
    "day_header": "Күн",
    "days": {
        "monday": "Дүйсенбі",
        "tuesday": "Сейсенбі",
        "wednesday": "Сәрсенбі",
        "thursday": "Бейсенбі",
        "friday": "Жұма",
    },
    "stream_types": {
        "lecture": "Дәріс",
        "practical": "Практика",
        "lab": "Зертхана",
    },
}

INSTRUCTOR_STRINGS_RUS = {
    "university": "Западно-Казахстанский инновационно-технологический университет",
    "schedule_title": "РАСПИСАНИЕ ПРЕПОДАВАТЕЛЯ",
    "legend_both": "Обе недели",
    "legend_odd": "Нечетная неделя",
    "legend_even": "Четная неделя",
    "day_header": "День",
    "days": {
        "monday": "Понедельник",
        "tuesday": "Вторник",
        "wednesday": "Среда",
        "thursday": "Четверг",
        "friday": "Пятница",
    },
    "stream_types": {
        "lecture": "Лекция",
        "practical": "Практика",
        "lab": "Лаборатория",
    },
}

# Days in order
DAYS_ORDER = ["monday", "tuesday", "wednesday", "thursday", "friday"]

# Fonts
FONT_TITLE = Font(name="Times New Roman", size=16, bold=True)
FONT_HEADER = Font(name="Times New Roman", size=12, bold=True)
FONT_INSTRUCTOR_NAME = Font(name="Times New Roman", size=14, bold=True)
FONT_DAY = Font(name="Times New Roman", size=11, bold=True)
FONT_CELL = Font(name="Times New Roman", size=9, bold=False)
FONT_LEGEND = Font(name="Times New Roman", size=10, bold=False)

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER_NOWRAP = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Borders
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Week type fill colors
FILL_BOTH_WEEKS = PatternFill(
    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
)
FILL_ODD_WEEK = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
FILL_EVEN_WEEK = PatternFill(
    start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"
)

# Column widths for instructor schedule
INSTRUCTOR_COLUMN_WIDTHS = {
    "A": 14.0,  # Day column
}
# Slot columns B through N (slots 1-13)
for i, col in enumerate("BCDEFGHIJKLMN"):
    INSTRUCTOR_COLUMN_WIDTHS[col] = 22.0


@dataclass
class InstructorGeneratorConfig:
    """Configuration for instructor schedule generation."""

    language: Literal["kaz", "rus"]


class InstructorScheduleExcelGenerator:
    """Generates Excel instructor schedule files from JSON data."""

    def __init__(self, config: InstructorGeneratorConfig):
        """Initialize generator with configuration.

        Args:
            config: Generator configuration with language.
        """
        self.config = config
        self.strings = (
            INSTRUCTOR_STRINGS_KAZ
            if config.language == "kaz"
            else INSTRUCTOR_STRINGS_RUS
        )

    def load_json(self, path: Path) -> dict:
        """Load schedule data from JSON file.

        Args:
            path: Path to JSON file.

        Returns:
            Parsed JSON data as dictionary.
        """
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def group_by_instructor(self, data: dict) -> dict[str, list[dict]]:
        """Group assignments by instructor name.

        Uses normalized names for grouping to handle inconsistent spacing
        in source data (e.g., "а.о. Name" vs "а.о.Name").

        Args:
            data: Full schedule data with assignments.

        Returns:
            Dictionary mapping normalized instructor names to their assignments.
        """
        assignments = data.get("assignments", [])
        by_instructor: dict[str, list[dict]] = {}

        for assignment in assignments:
            instructor = assignment.get("instructor", "Unknown")
            # Normalize name for consistent grouping
            normalized = self.normalize_instructor_name(instructor)
            if normalized not in by_instructor:
                by_instructor[normalized] = []
            by_instructor[normalized].append(assignment)

        return by_instructor

    def build_instructor_grid(
        self, assignments: list[dict]
    ) -> dict[str, dict[int, list[dict]]]:
        """Build 2D schedule grid for an instructor.

        Args:
            assignments: List of assignment dictionaries for one instructor.

        Returns:
            Grid structure: {day: {slot: [assignments]}}
            Multiple assignments possible per slot (odd/even weeks).
        """
        grid: dict[str, dict[int, list[dict]]] = {}

        for day in DAYS_ORDER:
            grid[day] = {}
            for slot in range(1, 14):  # Slots 1-13
                grid[day][slot] = []

        for assignment in assignments:
            day = assignment.get("day", "")
            slot = assignment.get("slot", 0)
            if day in grid and slot in grid[day]:
                grid[day][slot].append(assignment)

        return grid

    def format_cell_content(self, assignment: dict) -> str:
        """Format assignment for cell display.

        Args:
            assignment: Assignment dictionary.

        Returns:
            Formatted multi-line string for cell.
        """
        stream_type = assignment.get("stream_type", "lecture")
        subject = assignment.get("subject", "")

        # Uppercase for lectures
        if stream_type == "lecture":
            subject = subject.upper()

        # Get localized stream type
        stream_type_localized = self.strings["stream_types"].get(
            stream_type, stream_type
        )

        # Format groups
        groups = assignment.get("groups", [])
        groups_str = ", ".join(groups) if groups else ""

        # Room info
        room = assignment.get("room", "")
        room_address = assignment.get("room_address", "")

        lines = [
            subject,
            f"[{stream_type_localized}]",
            groups_str,
            f"{room}, {room_address}"
            if room and room_address
            else room or room_address,
        ]

        return "\n".join(line for line in lines if line)

    def get_cell_fill(self, assignments: list[dict]) -> PatternFill:
        """Determine background color based on week type.

        Args:
            assignments: List of assignments for a cell.

        Returns:
            PatternFill for the cell background.
        """
        if not assignments:
            return FILL_BOTH_WEEKS

        week_types = {a.get("week_type", "both") for a in assignments}

        if "both" in week_types:
            return FILL_BOTH_WEEKS
        elif week_types == {"odd"}:
            return FILL_ODD_WEEK
        elif week_types == {"even"}:
            return FILL_EVEN_WEEK
        elif "odd" in week_types and "even" in week_types:
            # Both odd and even specific assignments - use white
            return FILL_BOTH_WEEKS
        else:
            return FILL_BOTH_WEEKS

    @staticmethod
    def sanitize_sheet_name(name: str) -> str:
        """Sanitize sheet name by removing invalid characters.

        Excel sheet names cannot contain: / \\ * ? : [ ]

        Args:
            name: Original sheet name.

        Returns:
            Sanitized sheet name (max 31 chars).
        """
        # Normalize first (handles prefixes and spacing)
        cleaned = InstructorScheduleExcelGenerator.normalize_instructor_name(name)
        # Replace invalid characters
        invalid_chars = r"/\*?:[]"
        for char in invalid_chars:
            cleaned = cleaned.replace(char, "")
        return cleaned[:31]

    @staticmethod
    def normalize_instructor_name(name: str) -> str:
        """Normalize instructor name for grouping.

        Handles inconsistent spacing after prefixes like "а.о." vs "а.о. ".

        Args:
            name: Raw instructor name.

        Returns:
            Normalized instructor name for consistent grouping.
        """
        import re

        # Remove common prefixes (with or without trailing space)
        cleaned = name.replace("а.о.", "").replace("қ.проф.", "")
        # Normalize multiple spaces to single space
        cleaned = re.sub(r"\s+", " ", cleaned)
        return cleaned.strip()

    @staticmethod
    def clean_instructor_name(name: str) -> str:
        """Clean instructor name for display.

        Args:
            name: Raw instructor name.

        Returns:
            Cleaned instructor name.
        """
        return InstructorScheduleExcelGenerator.normalize_instructor_name(name)

    def create_workbook(self, data: dict) -> Workbook:
        """Create Excel workbook with one sheet per instructor.

        Args:
            data: Full schedule data.

        Returns:
            Populated Workbook object.
        """
        wb = Workbook()
        wb.remove(wb.active)

        by_instructor = self.group_by_instructor(data)

        if not by_instructor:
            ws = wb.create_sheet(title="Empty")
            return wb

        # Sort instructors alphabetically
        for instructor_name in sorted(by_instructor.keys()):
            assignments = by_instructor[instructor_name]
            grid = self.build_instructor_grid(assignments)

            # Create sheet
            sheet_name = self.sanitize_sheet_name(instructor_name)
            ws = wb.create_sheet(title=sheet_name)

            # Setup sheet
            self.setup_instructor_sheet(ws, instructor_name)

            # Fill schedule
            self.fill_schedule(ws, grid)

        if not wb.worksheets:
            wb.create_sheet(title="Empty")

        return wb

    def setup_instructor_sheet(self, ws, instructor_name: str) -> None:
        """Set up sheet structure with headers and formatting.

        Args:
            ws: Worksheet to set up.
            instructor_name: Name of the instructor.
        """
        # Set column widths
        for col, width in INSTRUCTOR_COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

        # Set row heights
        ws.row_dimensions[2].height = 25.0  # University name
        ws.row_dimensions[4].height = 25.0  # Schedule title
        ws.row_dimensions[5].height = 25.0  # Instructor name
        ws.row_dimensions[7].height = 20.0  # Legend
        ws.row_dimensions[9].height = 30.0  # Headers
        # Data rows (5 days * 1 row each = rows 10-14)
        for row in range(10, 15):
            ws.row_dimensions[row].height = 100.0

        # Merge cells for header
        ws.merge_cells("A2:N2")  # University name
        ws.merge_cells("A4:N4")  # Schedule title
        ws.merge_cells("A5:N5")  # Instructor name

        # Row 2: University name
        ws["A2"] = self.strings["university"]
        ws["A2"].font = FONT_TITLE
        ws["A2"].alignment = ALIGN_CENTER

        # Row 4: Schedule title
        ws["A4"] = self.strings["schedule_title"]
        ws["A4"].font = FONT_TITLE
        ws["A4"].alignment = ALIGN_CENTER

        # Row 5: Instructor name
        cleaned_name = self.clean_instructor_name(instructor_name)
        ws["A5"] = cleaned_name
        ws["A5"].font = FONT_INSTRUCTOR_NAME
        ws["A5"].alignment = ALIGN_CENTER

        # Row 7: Legend
        self.setup_legend(ws)

        # Row 9: Headers (Day | Slot 1 | Slot 2 | ... | Slot 13)
        self.setup_headers(ws)

        # Setup grid structure
        self.setup_grid(ws)

    def setup_legend(self, ws) -> None:
        """Set up the legend row with color boxes.

        Args:
            ws: Worksheet.
        """
        row = 7

        # Legend items: (column, fill, label)
        legend_items = [
            ("B", FILL_BOTH_WEEKS, self.strings["legend_both"]),
            ("D", FILL_ODD_WEEK, self.strings["legend_odd"]),
            ("F", FILL_EVEN_WEEK, self.strings["legend_even"]),
        ]

        for col, fill, label in legend_items:
            cell = ws[f"{col}{row}"]
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER

            # Label in next column
            next_col = chr(ord(col) + 1)
            label_cell = ws[f"{next_col}{row}"]
            label_cell.value = label
            label_cell.font = FONT_LEGEND
            label_cell.alignment = ALIGN_LEFT

    def setup_headers(self, ws) -> None:
        """Set up header row with day and slot columns.

        Args:
            ws: Worksheet.
        """
        row = 9

        # Day header (column A)
        ws[f"A{row}"] = self.strings["day_header"]
        ws[f"A{row}"].font = FONT_HEADER
        ws[f"A{row}"].alignment = ALIGN_CENTER
        ws[f"A{row}"].border = THIN_BORDER

        # Slot headers (columns B-N for slots 1-13)
        slot_cols = "BCDEFGHIJKLMN"
        for i, col in enumerate(slot_cols):
            slot_num = i + 1
            time_range = get_slot_time_range(slot_num)
            ws[f"{col}{row}"] = f"{slot_num}\n{time_range}"
            ws[f"{col}{row}"].font = FONT_HEADER
            ws[f"{col}{row}"].alignment = ALIGN_CENTER
            ws[f"{col}{row}"].border = THIN_BORDER

    def setup_grid(self, ws) -> None:
        """Set up the schedule grid with empty cells.

        Args:
            ws: Worksheet.
        """
        start_row = 10

        for i, day in enumerate(DAYS_ORDER):
            row = start_row + i

            # Day name (column A)
            ws[f"A{row}"] = self.strings["days"][day]
            ws[f"A{row}"].font = FONT_DAY
            ws[f"A{row}"].alignment = ALIGN_CENTER
            ws[f"A{row}"].border = THIN_BORDER

            # Slot cells (columns B-N)
            slot_cols = "BCDEFGHIJKLMN"
            for col in slot_cols:
                ws[f"{col}{row}"].border = THIN_BORDER
                ws[f"{col}{row}"].alignment = ALIGN_CENTER
                ws[f"{col}{row}"].font = FONT_CELL

    def fill_schedule(self, ws, grid: dict[str, dict[int, list[dict]]]) -> None:
        """Fill schedule cells with assignment data.

        Args:
            ws: Worksheet.
            grid: Schedule grid from build_instructor_grid.
        """
        start_row = 10
        slot_cols = "BCDEFGHIJKLMN"

        for i, day in enumerate(DAYS_ORDER):
            row = start_row + i

            if day not in grid:
                continue

            for slot_num, assignments in grid[day].items():
                if not assignments:
                    continue

                col = slot_cols[slot_num - 1]

                # Format cell content
                # If multiple assignments (odd/even), combine them
                if len(assignments) == 1:
                    content = self.format_cell_content(assignments[0])
                else:
                    # Multiple assignments for same slot (odd/even weeks)
                    contents = []
                    for a in assignments:
                        week_type = a.get("week_type", "both")
                        week_label = (
                            f"[{self.strings['legend_odd']}]"
                            if week_type == "odd"
                            else f"[{self.strings['legend_even']}]"
                            if week_type == "even"
                            else ""
                        )
                        a_content = self.format_cell_content(a)
                        if week_label:
                            contents.append(f"{week_label}\n{a_content}")
                        else:
                            contents.append(a_content)
                    content = "\n---\n".join(contents)

                ws[f"{col}{row}"] = content
                ws[f"{col}{row}"].fill = self.get_cell_fill(assignments)

    def save(self, wb: Workbook, output_path: Path) -> None:
        """Save workbook to file.

        Args:
            wb: Workbook to save.
            output_path: Output file path.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)


def generate_instructor_schedule_excel(
    input_path: Path,
    output_dir: Path,
    language: str | None = None,
) -> list[Path]:
    """Generate Excel instructor schedule files from JSON.

    Args:
        input_path: Path to schedule JSON file.
        output_dir: Output directory for Excel files.
        language: Language ('kaz' or 'rus'), or None for both.

    Returns:
        List of generated file paths.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    generated_files: list[Path] = []

    # Determine what to generate
    languages = [language] if language else ["kaz", "rus"]

    for lang in languages:
        config = InstructorGeneratorConfig(language=lang)  # type: ignore
        generator = InstructorScheduleExcelGenerator(config)

        # Load data
        data = generator.load_json(input_path)

        # Create and save workbook
        wb = generator.create_workbook(data)
        output_file = output_dir / f"instructor_schedules_{lang}.xlsx"
        generator.save(wb, output_file)
        generated_files.append(output_file)

    return generated_files
