"""Excel generator for unscheduled streams report (in Russian)."""

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

# Russian translations for UnscheduledReason values
UNSCHEDULED_REASON_RUS = {
    "instructor_conflict": "Конфликт расписания преподавателя",
    "group_conflict": "Конфликт расписания группы",
    "no_room_available": "Нет свободной аудитории",
    "instructor_unavailable": "Преподаватель недоступен",
    "no_consecutive_slots": "Нет последовательных слотов",
    "all_slots_exhausted": "Все слоты заняты",
    "building_gap_required": "Требуется перерыв между корпусами",
    "no_lecture_dependency": "Нет связанной лекции",
    "subject_daily_limit_exceeded": "Превышен дневной лимит предмета",
    "daily_load_exceeded": "Превышена дневная нагрузка группы",
    "max_windows_exceeded": "Превышено количество окон",
    "instructor_day_constraint": "Ограничение дня преподавателя",
    "no_parallel_subgroup": "Нет параллельной подгруппы",
    "subgroup_pairing_failed": "Не удалось сопоставить подгруппы",
}

# Shift translations
SHIFT_RUS = {
    "first": "1-я смена",
    "second": "2-я смена",
}

# Localization strings
STRINGS_RUS = {
    "title": "НЕРАСПРЕДЕЛЕННЫЕ ЗАНЯТИЯ",
    "generated": "Дата создания:",
    "total_count": "Всего нераспределенных:",
    "by_reason": "Распределение по причинам:",
    "col_number": "№",
    "col_subject": "Предмет",
    "col_instructor": "Преподаватель",
    "col_groups": "Группы",
    "col_student_count": "Кол-во студ.",
    "col_shift": "Смена",
    "col_reason": "Причина",
    "col_details": "Подробности",
}

# Column widths
COLUMN_WIDTHS = {
    "A": 5.0,   # №
    "B": 40.0,  # Предмет
    "C": 25.0,  # Преподаватель
    "D": 25.0,  # Группы
    "E": 12.0,  # Кол-во студ.
    "F": 10.0,  # Смена
    "G": 35.0,  # Причина
    "H": 55.0,  # Подробности
}

# Fonts
FONT_TITLE = Font(name="Times New Roman", size=16, bold=True)
FONT_HEADER = Font(name="Times New Roman", size=12, bold=True)
FONT_SUMMARY = Font(name="Times New Roman", size=11, bold=False)
FONT_SUMMARY_BOLD = Font(name="Times New Roman", size=11, bold=True)
FONT_CELL = Font(name="Times New Roman", size=11, bold=False)

# Alignments
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Borders
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@dataclass
class UnscheduledGeneratorConfig:
    """Configuration for unscheduled streams report."""

    pass  # No config needed for now, but allows future extension


class UnscheduledExcelGenerator:
    """Generates Excel report of unscheduled streams in Russian."""

    def __init__(self):
        """Initialize generator with Russian localization."""
        self.strings = STRINGS_RUS

    def load_json(self, path: Path) -> dict:
        """Load schedule data from JSON file.

        Args:
            path: Path to JSON file.

        Returns:
            Parsed JSON data as dictionary.
        """
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def translate_reason(self, reason: str) -> str:
        """Translate reason code to Russian.

        Args:
            reason: English reason code.

        Returns:
            Russian translation or original if not found.
        """
        return UNSCHEDULED_REASON_RUS.get(reason, reason)

    def translate_shift(self, shift: str) -> str:
        """Translate shift to Russian.

        Args:
            shift: English shift name.

        Returns:
            Russian translation or original if not found.
        """
        return SHIFT_RUS.get(shift, shift)

    def sort_streams(self, streams: list[dict]) -> list[dict]:
        """Sort streams by reason, instructor, subject.

        Args:
            streams: List of unscheduled stream dictionaries.

        Returns:
            Sorted list of streams.
        """
        return sorted(
            streams,
            key=lambda s: (
                s.get("reason", ""),
                s.get("instructor", ""),
                s.get("subject", ""),
            ),
        )

    def count_by_reason(self, streams: list[dict]) -> dict[str, int]:
        """Count streams by reason.

        Args:
            streams: List of unscheduled stream dictionaries.

        Returns:
            Dictionary mapping reason to count.
        """
        counts: dict[str, int] = {}
        for stream in streams:
            reason = stream.get("reason", "unknown")
            counts[reason] = counts.get(reason, 0) + 1
        return counts

    def setup_header(self, ws, total_count: int) -> int:
        """Set up header section with title and date.

        Args:
            ws: Worksheet to set up.
            total_count: Total number of unscheduled streams.

        Returns:
            Next row number after header.
        """
        # Row 1: Title
        ws.merge_cells("A1:H1")
        ws["A1"] = self.strings["title"]
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = ALIGN_CENTER

        # Row 3: Generated date
        ws["A3"] = self.strings["generated"]
        ws["A3"].font = FONT_SUMMARY_BOLD
        ws["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
        ws["B3"].font = FONT_SUMMARY

        # Row 4: Total count
        ws["A4"] = self.strings["total_count"]
        ws["A4"].font = FONT_SUMMARY_BOLD
        ws["B4"] = total_count
        ws["B4"].font = FONT_SUMMARY

        return 6  # Next row for summary

    def setup_summary(self, ws, streams: list[dict], start_row: int) -> int:
        """Set up summary section with breakdown by reason.

        Args:
            ws: Worksheet.
            streams: List of unscheduled stream dictionaries.
            start_row: Row to start summary.

        Returns:
            Next row number after summary.
        """
        counts = self.count_by_reason(streams)

        # Summary header
        ws[f"A{start_row}"] = self.strings["by_reason"]
        ws[f"A{start_row}"].font = FONT_SUMMARY_BOLD

        # Reason breakdown (mini-table)
        row = start_row + 1
        for reason, count in sorted(counts.items(), key=lambda x: -x[1]):
            ws[f"A{row}"] = f"  • {self.translate_reason(reason)}"
            ws[f"A{row}"].font = FONT_SUMMARY
            ws[f"B{row}"] = count
            ws[f"B{row}"].font = FONT_SUMMARY
            row += 1

        return row + 1  # Add one empty row before table

    def setup_table(self, ws, streams: list[dict], start_row: int) -> None:
        """Set up the main data table.

        Args:
            ws: Worksheet.
            streams: List of unscheduled stream dictionaries (already sorted).
            start_row: Row to start table.
        """
        # Table header
        headers = [
            ("A", self.strings["col_number"]),
            ("B", self.strings["col_subject"]),
            ("C", self.strings["col_instructor"]),
            ("D", self.strings["col_groups"]),
            ("E", self.strings["col_student_count"]),
            ("F", self.strings["col_shift"]),
            ("G", self.strings["col_reason"]),
            ("H", self.strings["col_details"]),
        ]

        for col, header_text in headers:
            cell = ws[f"{col}{start_row}"]
            cell.value = header_text
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

        # Set row height for header
        ws.row_dimensions[start_row].height = 25.0

        # Data rows
        for i, stream in enumerate(streams, 1):
            row = start_row + i

            # №
            ws[f"A{row}"] = i
            ws[f"A{row}"].font = FONT_CELL
            ws[f"A{row}"].alignment = ALIGN_CENTER
            ws[f"A{row}"].border = THIN_BORDER

            # Предмет
            ws[f"B{row}"] = stream.get("subject", "")
            ws[f"B{row}"].font = FONT_CELL
            ws[f"B{row}"].alignment = ALIGN_LEFT
            ws[f"B{row}"].border = THIN_BORDER

            # Преподаватель
            ws[f"C{row}"] = stream.get("instructor", "")
            ws[f"C{row}"].font = FONT_CELL
            ws[f"C{row}"].alignment = ALIGN_LEFT
            ws[f"C{row}"].border = THIN_BORDER

            # Группы
            groups = stream.get("groups", [])
            groups_text = ", ".join(groups) if isinstance(groups, list) else str(groups)
            ws[f"D{row}"] = groups_text
            ws[f"D{row}"].font = FONT_CELL
            ws[f"D{row}"].alignment = ALIGN_LEFT
            ws[f"D{row}"].border = THIN_BORDER

            # Кол-во студ.
            ws[f"E{row}"] = stream.get("student_count", 0)
            ws[f"E{row}"].font = FONT_CELL
            ws[f"E{row}"].alignment = ALIGN_CENTER
            ws[f"E{row}"].border = THIN_BORDER

            # Смена
            shift = stream.get("shift", "")
            ws[f"F{row}"] = self.translate_shift(shift)
            ws[f"F{row}"].font = FONT_CELL
            ws[f"F{row}"].alignment = ALIGN_CENTER
            ws[f"F{row}"].border = THIN_BORDER

            # Причина
            reason = stream.get("reason", "")
            ws[f"G{row}"] = self.translate_reason(reason)
            ws[f"G{row}"].font = FONT_CELL
            ws[f"G{row}"].alignment = ALIGN_LEFT
            ws[f"G{row}"].border = THIN_BORDER

            # Подробности
            ws[f"H{row}"] = stream.get("details", "")
            ws[f"H{row}"].font = FONT_CELL
            ws[f"H{row}"].alignment = ALIGN_LEFT_TOP
            ws[f"H{row}"].border = THIN_BORDER

            # Set row height for data rows
            ws.row_dimensions[row].height = 45.0

    def create_workbook(self, unscheduled_streams: list[dict]) -> Workbook:
        """Create Excel workbook with unscheduled streams report.

        Args:
            unscheduled_streams: List of unscheduled stream dictionaries.

        Returns:
            Populated Workbook object.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Нераспределенные"

        # Set column widths
        for col, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

        # Sort streams
        sorted_streams = self.sort_streams(unscheduled_streams)

        # Setup sections
        next_row = self.setup_header(ws, len(unscheduled_streams))
        next_row = self.setup_summary(ws, unscheduled_streams, next_row)
        self.setup_table(ws, sorted_streams, next_row)

        return wb

    def save(self, wb: Workbook, output_path: Path) -> None:
        """Save workbook to file.

        Args:
            wb: Workbook to save.
            output_path: Output file path.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)


def generate_unscheduled_excel(input_path: Path, output_path: Path) -> Path:
    """Generate Excel report of unscheduled streams.

    Args:
        input_path: Path to schedule JSON file.
        output_path: Output Excel file path.

    Returns:
        Path to generated file.
    """
    generator = UnscheduledExcelGenerator()

    # Load data
    data = generator.load_json(input_path)
    unscheduled_streams = data.get("unscheduled_streams", [])

    # Create and save workbook
    wb = generator.create_workbook(unscheduled_streams)
    generator.save(wb, output_path)

    return output_path
