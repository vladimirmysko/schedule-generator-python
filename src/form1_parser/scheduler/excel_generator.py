"""Excel schedule generator.

Generates formatted Excel schedule files from schedule results.
"""

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .utils import clean_instructor_name, parse_group_year

# Days order for schedule
DAYS_ORDER = ["monday", "tuesday", "wednesday", "thursday", "friday"]

# Row ranges for each day (7 slots per day)
DAY_ROW_RANGES = {
    "monday": (10, 16),
    "tuesday": (17, 23),
    "wednesday": (24, 30),
    "thursday": (31, 37),
    "friday": (38, 44),
}

# Maximum groups per sheet
MAX_GROUPS_PER_SHEET = 3


@dataclass
class GeneratorConfig:
    """Configuration for schedule generation."""

    language: str = "kaz"  # kaz or rus
    year: int = 1  # 1-5
    week_type: str = "both"  # odd, even, or both


# Localized strings
STRINGS = {
    "kaz": {
        "university": "БАТЫС ҚАЗАҚСТАН ИННОВАЦИЯЛЫҚ-ТЕХНОЛОГИЯЛЫҚ УНИВЕРСИТЕТІ",
        "schedule_title": "САБАҚ КЕСТЕСІ",
        "days": {
            "monday": "Дүйсенбі",
            "tuesday": "Сейсенбі",
            "wednesday": "Сәрсенбі",
            "thursday": "Бейсенбі",
            "friday": "Жұма",
        },
        "headers": {
            "day": "Күн",
            "time": "Уақыт",
            "slot": "№",
        },
        "course": "курс",
        "week_odd": "ТАҚ",
        "week_even": "ЖҰП",
    },
    "rus": {
        "university": "ЗАПАДНО-КАЗАХСТАНСКИЙ ИННОВАЦИОННО-ТЕХНОЛОГИЧЕСКИЙ УНИВЕРСИТЕТ",
        "schedule_title": "РАСПИСАНИЕ ЗАНЯТИЙ",
        "days": {
            "monday": "Понедельник",
            "tuesday": "Вторник",
            "wednesday": "Среда",
            "thursday": "Четверг",
            "friday": "Пятница",
        },
        "headers": {
            "day": "День",
            "time": "Время",
            "slot": "№",
        },
        "course": "курс",
        "week_odd": "НЕЧ",
        "week_even": "ЧЕТ",
    },
}

# Time slots
SLOT_TIMES = {
    1: "09:00-09:50",
    2: "10:00-10:50",
    3: "11:00-11:50",
    4: "12:00-12:50",
    5: "13:00-13:50",
    6: "14:00-14:50",
    7: "15:00-15:50",
    8: "16:00-16:50",
    9: "17:00-17:50",
    10: "18:00-18:50",
    11: "19:00-19:50",
    12: "20:00-20:50",
    13: "21:00-21:50",
}


class ScheduleExcelGenerator:
    """Generates Excel schedule files."""

    def __init__(self, config: GeneratorConfig):
        self.config = config
        self.strings = STRINGS[config.language]

    @staticmethod
    def get_year_from_group(group_name: str) -> int:
        """Extract year from group code."""
        year = parse_group_year(group_name)
        return year if year is not None else 0

    @staticmethod
    def is_russian_group(group_name: str) -> bool:
        """
        Determine if group is Russian language based on second digit.

        Odd (1,3,5,7,9) = Kazakh
        Even (2,4,6,8,0) = Russian
        """
        match = re.search(r"-\d(\d)", group_name)
        if match:
            second_digit = int(match.group(1))
            return second_digit % 2 == 0
        return False

    def filter_assignments(
        self, schedule_data: dict[str, Any]
    ) -> tuple[list[dict], set[str]]:
        """
        Filter assignments by year, language, and week type.

        Returns (filtered_assignments, matching_groups).
        """
        assignments = schedule_data.get("assignments", [])
        is_russian = self.config.language == "rus"
        target_year = self.config.year
        week_type = self.config.week_type

        filtered = []
        groups_found = set()

        for assignment in assignments:
            # Check week type
            a_week = assignment.get("week_type", "both")
            if week_type == "odd" and a_week not in ("odd", "both"):
                continue
            if week_type == "even" and a_week not in ("even", "both"):
                continue

            # Check groups match year and language
            matching_groups = []
            for group in assignment.get("groups", []):
                year = self.get_year_from_group(group)
                if year != target_year:
                    continue
                if is_russian != self.is_russian_group(group):
                    continue
                matching_groups.append(group)
                groups_found.add(group)

            if matching_groups:
                filtered.append(assignment)

        return filtered, groups_found

    def group_into_sheets(self, groups: list[str] | set[str]) -> list[list[str]]:
        """
        Group groups into sheets (max 3 per sheet).

        Returns list of group lists, one per sheet.
        """
        sorted_groups = sorted(groups)
        sheets = []
        for i in range(0, len(sorted_groups), MAX_GROUPS_PER_SHEET):
            sheets.append(sorted_groups[i : i + MAX_GROUPS_PER_SHEET])
        return sheets

    def build_schedule_grid(
        self, assignments: list[dict], groups: list[str]
    ) -> dict[str, dict[int, dict[str, dict | None]]]:
        """
        Build schedule grid from assignments.

        Returns: {day: {slot: {group: assignment}}}
        """
        grid: dict[str, dict[int, dict[str, dict | None]]] = {}

        # Initialize grid
        for day in DAYS_ORDER:
            grid[day] = {}
            for slot in range(1, 14):  # Slots 1-13
                grid[day][slot] = {group: None for group in groups}

        # Fill grid with assignments
        for assignment in assignments:
            day = assignment.get("day", "").lower()
            slot = assignment.get("slot", 0)
            if day not in grid or slot not in grid[day]:
                continue

            for group in assignment.get("groups", []):
                if group in grid[day][slot]:
                    grid[day][slot][group] = assignment

        return grid

    def format_cell_content(self, assignment: dict) -> str:
        """Format assignment for cell display."""
        subject = assignment.get("subject", "").upper()
        instructor = clean_instructor_name(assignment.get("instructor", ""))
        room = assignment.get("room", "")
        room_address = assignment.get("room_address", "")

        lines = [subject, instructor, f"{room}, {room_address}"]
        return "\n".join(lines)

    def create_workbook(
        self, assignments: list[dict], groups: set[str] | list[str]
    ) -> Workbook:
        """Create Excel workbook with schedule."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        if not groups:
            ws = wb.create_sheet("Empty")
            ws["A1"] = "No groups found for this configuration"
            return wb

        sheets = self.group_into_sheets(groups)

        for sheet_idx, sheet_groups in enumerate(sheets):
            sheet_name = f"Sheet{sheet_idx + 1}"
            ws = wb.create_sheet(sheet_name)
            self._setup_sheet(ws, sheet_groups, assignments)

        return wb

    def _setup_sheet(
        self, ws, groups: list[str], assignments: list[dict]
    ) -> None:
        """Set up a single sheet with schedule data."""
        strings = self.strings

        # University name
        ws["A2"] = strings["university"]
        ws.merge_cells("A2:F2")

        # Schedule title
        ws["A6"] = strings["schedule_title"]
        ws.merge_cells("A6:F6")

        # Course/Year
        ws["A8"] = f"{self.config.year} {strings['course']}"

        # Headers
        ws["A9"] = strings["headers"]["day"]
        ws["B9"] = strings["headers"]["time"]
        ws["C9"] = strings["headers"]["slot"]

        # Group headers
        for i, group in enumerate(groups):
            col = get_column_letter(4 + i)
            ws[f"{col}9"] = group

        # Build grid
        grid = self.build_schedule_grid(assignments, groups)

        # Fill schedule
        for day_idx, day in enumerate(DAYS_ORDER):
            start_row, end_row = DAY_ROW_RANGES[day]

            # Day name
            ws[f"A{start_row}"] = strings["days"][day]
            ws.merge_cells(f"A{start_row}:A{end_row}")

            # Slots
            for slot_offset in range(7):
                slot = slot_offset + 1
                row = start_row + slot_offset

                # Time
                ws[f"B{row}"] = SLOT_TIMES.get(slot, "")
                # Slot number
                ws[f"C{row}"] = slot

                # Group cells
                for i, group in enumerate(groups):
                    col = get_column_letter(4 + i)
                    assignment = grid[day][slot].get(group)
                    if assignment:
                        ws[f"{col}{row}"] = self.format_cell_content(assignment)

        # Apply styling
        self._apply_styling(ws, groups)

    def _apply_styling(self, ws, groups: list[str]) -> None:
        """Apply basic styling to worksheet."""
        # Set column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 4

        for i in range(len(groups)):
            col = get_column_letter(4 + i)
            ws.column_dimensions[col].width = 25

        # Center alignment for headers
        header_font = Font(bold=True)
        for cell in ws[9]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def save(self, output_path: Path, assignments: list[dict], groups: set[str]) -> None:
        """Generate and save Excel file."""
        wb = self.create_workbook(assignments, groups)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)


def generate_schedule_excel(
    input_path: Path,
    output_dir: Path,
    language: str | None = None,
    year: int | None = None,
    week_type: str | None = None,
) -> list[Path]:
    """
    Generate Excel schedule files from JSON schedule data.

    Args:
        input_path: Path to schedule JSON file
        output_dir: Directory for output files
        language: Filter by language (kaz/rus) or None for all
        year: Filter by year (1-5) or None for all
        week_type: Filter by week type (odd/even/both) or None for all

    Returns:
        List of generated file paths.
    """
    with open(input_path, encoding="utf-8") as f:
        schedule_data = json.load(f)

    output_dir.mkdir(parents=True, exist_ok=True)
    generated_files = []

    # Determine combinations to generate
    languages = [language] if language else ["kaz", "rus"]
    years = [year] if year else [1, 2, 3, 4, 5]
    week_types = [week_type] if week_type else ["odd", "even", "both"]

    for lang in languages:
        for y in years:
            for wt in week_types:
                config = GeneratorConfig(language=lang, year=y, week_type=wt)
                generator = ScheduleExcelGenerator(config)

                assignments, groups = generator.filter_assignments(schedule_data)
                if not groups:
                    continue

                filename = f"schedule_{lang}_{y}y_{wt}.xlsx"
                output_path = output_dir / filename
                generator.save(output_path, assignments, groups)
                generated_files.append(output_path)

    return generated_files
