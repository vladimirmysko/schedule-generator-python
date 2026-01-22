"""Tests for Excel schedule generator."""

import json
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from form1_parser.scheduler.excel_generator import (
    DAYS_ORDER,
    GeneratorConfig,
    ScheduleExcelGenerator,
    analyze_group_structure,
    generate_schedule_excel,
)


# Sample test data
# Language is determined by second digit: odd → Kazakh, even → Russian
SAMPLE_ASSIGNMENTS = [
    {
        "stream_id": "test_stream_1",
        "subject": "Test Subject 1",
        "instructor": "а.о.Test Instructor",
        "groups": ["АРХ-11 О", "АРХ-13 О"],  # Kazakh (odd: 1, 3)
        "student_count": 50,
        "day": "monday",
        "slot": 1,
        "time": "09:00-09:50",
        "room": "A-101",
        "room_address": "Test Address 1",
        "week_type": "both",
    },
    {
        "stream_id": "test_stream_2",
        "subject": "Test Subject 2",
        "instructor": "қ.проф.Another Instructor",
        "groups": ["АРХ-21 О", "АРХ-23 О"],  # Kazakh (odd: 1, 3)
        "student_count": 40,
        "day": "tuesday",
        "slot": 2,
        "time": "10:00-10:50",
        "room": "B-202",
        "room_address": "Test Address 2",
        "week_type": "odd",
    },
    {
        "stream_id": "test_stream_3",
        "subject": "Test Subject 3",
        "instructor": "а.о.Third Instructor",
        "groups": ["ЮР-12 О", "ЮР-14 О"],  # Russian (even: 2, 4)
        "student_count": 30,
        "day": "wednesday",
        "slot": 3,
        "time": "11:00-11:50",
        "room": "C-303",
        "room_address": "Test Address 3",
        "week_type": "even",
    },
    {
        "stream_id": "test_stream_4",
        "subject": "Test Subject 4",
        "instructor": "Fourth Instructor",
        "groups": ["АРХ-11 О", "АРХ-11А О", "АРХ-13 О", "АРХ-13А О"],  # Kazakh
        "student_count": 80,
        "day": "monday",
        "slot": 2,
        "time": "10:00-10:50",
        "room": "D-404",
        "room_address": "Test Address 4",
        "week_type": "both",
    },
]

SAMPLE_SCHEDULE_DATA = {
    "generation_date": "2025-01-20T10:00:00",
    "stage": 1,
    "total_assigned": 4,
    "total_unscheduled": 0,
    "assignments": SAMPLE_ASSIGNMENTS,
    "unscheduled_stream_ids": [],
    "statistics": {},
}


class TestGetYearFromGroup:
    """Tests for year extraction from group codes."""

    def test_year_1(self):
        assert ScheduleExcelGenerator.get_year_from_group("АРХ-11 О") == 1
        assert ScheduleExcelGenerator.get_year_from_group("ЮР-15 О") == 1

    def test_year_2(self):
        assert ScheduleExcelGenerator.get_year_from_group("АРХ-21 О") == 2
        assert ScheduleExcelGenerator.get_year_from_group("ЮР-23 О") == 2

    def test_year_3(self):
        assert ScheduleExcelGenerator.get_year_from_group("АРХ-31 О") == 3

    def test_year_4(self):
        assert ScheduleExcelGenerator.get_year_from_group("АРХ-41 О") == 4

    def test_no_year(self):
        assert ScheduleExcelGenerator.get_year_from_group("INVALID") == 0

    def test_group_with_suffix(self):
        assert ScheduleExcelGenerator.get_year_from_group("АРХ-11А О") == 1
        assert ScheduleExcelGenerator.get_year_from_group("ЮР-22 О /г/") == 2


class TestIsRussianGroup:
    """Tests for Russian language detection.

    Language is determined by second digit:
    - Odd (1,3,5,7,9) → Kazakh
    - Even (2,4,6,8,0) → Russian
    """

    def test_kazakh_group_odd_second_digit(self):
        # Second digit 1 (odd) → Kazakh
        assert not ScheduleExcelGenerator.is_russian_group("АРХ-11 О")
        assert not ScheduleExcelGenerator.is_russian_group("АУ-31 О")
        # Second digit 3 (odd) → Kazakh
        assert not ScheduleExcelGenerator.is_russian_group("ЮР-13 О")
        assert not ScheduleExcelGenerator.is_russian_group("ЮР-33 О /у/")
        # Second digit 5 (odd) → Kazakh
        assert not ScheduleExcelGenerator.is_russian_group("ЮР-15 О")

    def test_russian_group_even_second_digit(self):
        # Second digit 2 (even) → Russian
        assert ScheduleExcelGenerator.is_russian_group("СТР-22 О")
        assert ScheduleExcelGenerator.is_russian_group("АРХ-22 О")
        assert ScheduleExcelGenerator.is_russian_group("ЮР-12 О")
        # Second digit 4 (even) → Russian
        assert ScheduleExcelGenerator.is_russian_group("АРХ-14 О")
        # Second digit 0 (even) → Russian
        assert ScheduleExcelGenerator.is_russian_group("ЮР-20 О")

    def test_group_with_markers_still_uses_digit(self):
        # Markers don't matter, only second digit
        assert not ScheduleExcelGenerator.is_russian_group("ЮР-11 О /г/")  # odd → Kazakh
        assert ScheduleExcelGenerator.is_russian_group("ЮР-12 О /г/")  # even → Russian


class TestFilterAssignments:
    """Tests for assignment filtering."""

    def test_filter_by_year_1(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            SAMPLE_SCHEDULE_DATA
        )

        # Should include year 1 Kazakh groups (odd second digit)
        assert "АРХ-11 О" in base_groups
        assert "АРХ-13 О" in base_groups
        # Should exclude year 2
        assert "АРХ-21 О" not in base_groups
        # Should exclude Russian groups (even second digit)
        assert "ЮР-12 О" not in base_groups

    def test_filter_by_year_2(self):
        config = GeneratorConfig(language="kaz", year=2, week_type="odd")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            SAMPLE_SCHEDULE_DATA
        )

        assert "АРХ-21 О" in base_groups
        assert "АРХ-23 О" in base_groups
        assert "АРХ-11 О" not in base_groups

    def test_filter_russian_groups(self):
        config = GeneratorConfig(language="rus", year=1, week_type="even")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            SAMPLE_SCHEDULE_DATA
        )

        # Russian groups have even second digit (2, 4)
        assert "ЮР-12 О" in base_groups
        assert "ЮР-14 О" in base_groups
        # Kazakh groups (odd second digit) should be excluded
        assert "АРХ-11 О" not in base_groups

    def test_filter_by_week_type_odd(self):
        config = GeneratorConfig(language="kaz", year=2, week_type="odd")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            SAMPLE_SCHEDULE_DATA
        )

        # Should include odd and both week types
        assert len(assignments) > 0

    def test_filter_by_week_type_even(self):
        config = GeneratorConfig(language="rus", year=1, week_type="even")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            SAMPLE_SCHEDULE_DATA
        )

        # Should include even and both week types
        assert len(assignments) > 0


class TestGroupIntoSheets:
    """Tests for grouping into sheets."""

    def test_single_group(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        groups = ["АРХ-11 О"]
        sheets = generator.group_into_sheets(groups)

        assert len(sheets) == 1
        assert sheets[0] == ["АРХ-11 О"]

    def test_three_groups(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        groups = ["АРХ-11 О", "АРХ-13 О", "АРХ-15 О"]
        sheets = generator.group_into_sheets(groups)

        assert len(sheets) == 1
        assert len(sheets[0]) == 3

    def test_four_groups(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        groups = ["АРХ-11 О", "АРХ-13 О", "АРХ-15 О", "АРХ-17 О"]
        sheets = generator.group_into_sheets(groups)

        # 4 groups should fit in 1 sheet (max 5 per sheet)
        assert len(sheets) == 1
        assert len(sheets[0]) == 4

    def test_five_groups(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        groups = ["G1", "G2", "G3", "G4", "G5"]
        sheets = generator.group_into_sheets(groups)

        # 5 groups should fit in 1 sheet (max 5 per sheet)
        assert len(sheets) == 1
        assert len(sheets[0]) == 5

    def test_six_groups(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        groups = ["G1", "G2", "G3", "G4", "G5", "G6"]
        sheets = generator.group_into_sheets(groups)

        # 6 groups should split into 2 sheets (5 + 1)
        assert len(sheets) == 2
        assert len(sheets[0]) == 5
        assert len(sheets[1]) == 1


class TestBuildScheduleGrid:
    """Tests for schedule grid construction."""

    def test_grid_structure(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        groups = ["АРХ-11 О", "АРХ-13 О"]
        group_structure = analyze_group_structure(groups)
        grid = generator.build_schedule_grid(SAMPLE_ASSIGNMENTS, groups, group_structure)

        # Check all days are present
        for day in DAYS_ORDER:
            assert day in grid

        # Check slots are present
        for slot in range(1, 8):
            assert slot in grid["monday"]

        # Check groups are in each slot (now with sub1/sub2 structure)
        for group in groups:
            assert group in grid["monday"][1]
            assert "sub1" in grid["monday"][1][group]
            assert "sub2" in grid["monday"][1][group]

    def test_grid_assignment_placement(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        groups = ["АРХ-11 О", "АРХ-13 О"]
        group_structure = analyze_group_structure(groups)
        grid = generator.build_schedule_grid(SAMPLE_ASSIGNMENTS, groups, group_structure)

        # Check that test_stream_1 is placed in monday slot 1 (in sub1)
        assert grid["monday"][1]["АРХ-11 О"]["sub1"] is not None
        assert grid["monday"][1]["АРХ-13 О"]["sub1"] is not None
        assert grid["monday"][1]["АРХ-11 О"]["sub1"]["subject"] == "Test Subject 1"


class TestFormatCellContent:
    """Tests for cell content formatting."""

    def test_format_single_assignment(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignment = {
            "subject": "Test Subject",
            "instructor": "Test Instructor",
            "room": "A-101",
            "room_address": "Test Address",
            "stream_type": "lecture",
        }
        # Single group (no subgroups)
        content = generator.format_cell_content(assignment, None, False)

        assert "TEST SUBJECT" in content
        assert "Test Instructor" in content
        assert "A-101, Test Address" in content

    def test_format_cleans_instructor_titles(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignment = {
            "subject": "Subject",
            "instructor": "а.о.Instructor Name",
            "room": "A-101",
            "room_address": "Address",
            "stream_type": "lecture",
        }
        content = generator.format_cell_content(assignment, None, False)

        assert "а.о." not in content
        assert "Instructor Name" in content

    def test_format_cleans_professor_title(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignment = {
            "subject": "Subject",
            "instructor": "қ.проф.Professor Name",
            "room": "A-101",
            "room_address": "Address",
            "stream_type": "lecture",
        }
        content = generator.format_cell_content(assignment, None, False)

        assert "қ.проф." not in content
        assert "Professor Name" in content

    def test_format_side_by_side_both_subgroups(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        sub1_assignment = {
            "stream_id": "stream1",
            "subject": "Subject 1",
            "instructor": "Instructor 1",
            "room": "A-101",
            "room_address": "Address 1",
            "stream_type": "practical",
        }
        sub2_assignment = {
            "stream_id": "stream2",
            "subject": "Subject 2",
            "instructor": "Instructor 2",
            "room": "B-202",
            "room_address": "Address 2",
            "stream_type": "practical",
        }
        content = generator.format_cell_content(sub1_assignment, sub2_assignment, True)

        # Should have separator
        assert "-" in content
        assert "Subject 1" in content
        assert "Subject 2" in content

    def test_format_side_by_side_same_class(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignment = {
            "stream_id": "same_stream",
            "subject": "Same Subject",
            "instructor": "Same Instructor",
            "room": "A-101",
            "room_address": "Address",
            "stream_type": "practical",
        }
        # Both subgroups have same class (same stream_id)
        content = generator.format_cell_content(assignment, assignment, True)

        # Should show single content (no split)
        assert " - " not in content
        assert "Same Subject" in content

    def test_format_only_sub1_no_separator(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        sub1_assignment = {
            "stream_id": "stream1",
            "subject": "Subject 1",
            "instructor": "Instructor 1",
            "room": "A-101",
            "room_address": "Address",
            "stream_type": "practical",
        }
        content = generator.format_cell_content(sub1_assignment, None, True)

        # Whole group lesson - no separator
        assert "\n-\n" not in content
        assert "Subject 1" in content

    def test_format_only_sub2_no_separator(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        sub2_assignment = {
            "stream_id": "stream2",
            "subject": "Subject 2",
            "instructor": "Instructor 2",
            "room": "B-202",
            "room_address": "Address",
            "stream_type": "practical",
        }
        content = generator.format_cell_content(None, sub2_assignment, True)

        # Whole group lesson - no separator
        assert "\n-\n" not in content
        assert "Subject 2" in content


class TestCreateWorkbook:
    """Tests for workbook creation."""

    def test_creates_workbook(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            SAMPLE_SCHEDULE_DATA
        )
        wb = generator.create_workbook(assignments, base_groups, group_structure)

        assert wb is not None
        assert len(wb.worksheets) > 0

    def test_empty_groups_creates_empty_sheet(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        wb = generator.create_workbook([], [], {})

        assert wb is not None
        assert len(wb.worksheets) == 1
        assert "Empty" in wb.worksheets[0].title


class TestGenerateSingleFile:
    """Tests for single file generation."""

    def test_generate_single_file(self):
        with TemporaryDirectory() as tmpdir:
            # Write sample JSON
            input_path = Path(tmpdir) / "schedule.json"
            with open(input_path, "w", encoding="utf-8") as f:
                json.dump(SAMPLE_SCHEDULE_DATA, f)

            output_dir = Path(tmpdir) / "output"

            # Generate single file
            files = generate_schedule_excel(
                input_path=input_path,
                output_dir=output_dir,
                language="kaz",
                year=1,
                week_type="both",
            )

            assert len(files) == 1
            assert files[0].exists()
            assert "schedule_kaz_1y_both.xlsx" in str(files[0])

            # Verify workbook can be opened
            wb = load_workbook(files[0])
            assert len(wb.worksheets) > 0


class TestGenerateAllFiles:
    """Tests for generating all file combinations."""

    def test_generate_all_combinations(self):
        with TemporaryDirectory() as tmpdir:
            # Write sample JSON
            input_path = Path(tmpdir) / "schedule.json"
            with open(input_path, "w", encoding="utf-8") as f:
                json.dump(SAMPLE_SCHEDULE_DATA, f)

            output_dir = Path(tmpdir) / "output"

            # Generate all files (no filters)
            files = generate_schedule_excel(
                input_path=input_path,
                output_dir=output_dir,
            )

            # Should generate files for all combinations that have matching groups
            assert len(files) > 0

            # Check output directory exists
            assert output_dir.exists()

    def test_generate_by_language_filter(self):
        with TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "schedule.json"
            with open(input_path, "w", encoding="utf-8") as f:
                json.dump(SAMPLE_SCHEDULE_DATA, f)

            output_dir = Path(tmpdir) / "output"

            # Generate only Kazakh files
            files = generate_schedule_excel(
                input_path=input_path,
                output_dir=output_dir,
                language="kaz",
            )

            # All files should be Kazakh
            for file in files:
                assert "_kaz_" in str(file)
                assert "_rus_" not in str(file)


class TestSheetLayout:
    """Tests for sheet layout and formatting."""

    def test_sheet_has_correct_structure(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            SAMPLE_SCHEDULE_DATA
        )
        wb = generator.create_workbook(assignments, base_groups, group_structure)

        ws = wb.worksheets[0]

        # Check university name is in A2
        assert ws["A2"].value is not None

        # Check schedule title is in A6
        assert ws["A6"].value is not None

        # Check course/year is in A8
        assert ws["A8"].value is not None
        assert "1 курс" in str(ws["A8"].value)

        # Check headers in row 9
        assert ws["A9"].value is not None  # Day header
        assert ws["B9"].value is not None  # Time header
        assert ws["C9"].value == "№"  # Slot number header

    def test_sheet_has_day_names(self):
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            SAMPLE_SCHEDULE_DATA
        )
        wb = generator.create_workbook(assignments, base_groups, group_structure)

        ws = wb.worksheets[0]

        # Check day names are present in the sheet (rows are dynamic based on slots)
        strings = generator.strings

        # Collect all day name values from column A (starting from row 10)
        day_names_found = []
        for row in range(10, 100):  # Check rows 10-99
            cell_value = ws[f"A{row}"].value
            if cell_value and cell_value in strings["days"].values():
                day_names_found.append(cell_value)

        # Verify all day names are present in correct order
        expected_days = [
            strings["days"]["monday"],
            strings["days"]["tuesday"],
            strings["days"]["wednesday"],
            strings["days"]["thursday"],
            strings["days"]["friday"],
        ]
        assert day_names_found == expected_days


class TestAnalyzeGroupStructure:
    """Tests for group structure analysis."""

    def test_no_subgroups(self):
        groups = ["АРХ-11 О", "АРХ-13 О"]
        structure = analyze_group_structure(groups)

        assert len(structure) == 2
        assert "АРХ-11 О" in structure
        assert "АРХ-13 О" in structure
        assert structure["АРХ-11 О"]["has_subgroups"] is False
        assert structure["АРХ-13 О"]["has_subgroups"] is False
        assert structure["АРХ-11 О"]["subgroup_1"] == "АРХ-11 О"
        assert structure["АРХ-11 О"]["subgroup_2"] is None

    def test_with_subgroups(self):
        groups = ["СТР-11 О /1/", "СТР-11 О /2/", "АРХ-11 О"]
        structure = analyze_group_structure(groups)

        assert len(structure) == 2  # Base groups: СТР-11 О, АРХ-11 О
        assert "СТР-11 О" in structure
        assert "АРХ-11 О" in structure
        assert structure["СТР-11 О"]["has_subgroups"] is True
        assert structure["СТР-11 О"]["subgroup_1"] == "СТР-11 О /1/"
        assert structure["СТР-11 О"]["subgroup_2"] == "СТР-11 О /2/"
        assert structure["АРХ-11 О"]["has_subgroups"] is False

    def test_only_subgroup_1(self):
        groups = ["СТР-11 О /1/"]
        structure = analyze_group_structure(groups)

        assert len(structure) == 1
        assert "СТР-11 О" in structure
        assert structure["СТР-11 О"]["has_subgroups"] is True
        assert structure["СТР-11 О"]["subgroup_1"] == "СТР-11 О /1/"
        assert structure["СТР-11 О"]["subgroup_2"] is None

    def test_only_subgroup_2(self):
        groups = ["СТР-11 О /2/"]
        structure = analyze_group_structure(groups)

        assert len(structure) == 1
        assert "СТР-11 О" in structure
        assert structure["СТР-11 О"]["has_subgroups"] is True
        assert structure["СТР-11 О"]["subgroup_1"] is None
        assert structure["СТР-11 О"]["subgroup_2"] == "СТР-11 О /2/"

    def test_mixed_groups(self):
        groups = ["АРХ-11 О", "СТР-11 О /1/", "СТР-11 О /2/", "ЮР-11 О"]
        structure = analyze_group_structure(groups)

        assert len(structure) == 3
        assert structure["АРХ-11 О"]["has_subgroups"] is False
        assert structure["СТР-11 О"]["has_subgroups"] is True
        assert structure["ЮР-11 О"]["has_subgroups"] is False


class TestSubgroupDisplay:
    """Tests for subgroup display in Excel."""

    def test_subgroups_combined_in_filter(self):
        """Test that filter_assignments returns base groups, not subgroup variants."""
        # Create sample data with subgroups
        subgroup_data = {
            "assignments": [
                {
                    "stream_id": "stream1",
                    "subject": "Шетел тілі",
                    "instructor": "Instructor 1",
                    "groups": ["СТР-11 О /1/"],
                    "day": "monday",
                    "slot": 1,
                    "time": "09:00-09:50",
                    "room": "A-101",
                    "room_address": "Address",
                    "week_type": "both",
                },
                {
                    "stream_id": "stream2",
                    "subject": "Орыс тілі",
                    "instructor": "Instructor 2",
                    "groups": ["СТР-11 О /2/"],
                    "day": "monday",
                    "slot": 1,
                    "time": "09:00-09:50",
                    "room": "B-202",
                    "room_address": "Address",
                    "week_type": "both",
                },
            ]
        }

        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)
        assignments, base_groups, group_structure = generator.filter_assignments(
            subgroup_data
        )

        # Should have 1 base group (СТР-11 О), not 2 subgroup variants
        assert len(base_groups) == 1
        assert "СТР-11 О" in base_groups
        assert "СТР-11 О /1/" not in base_groups
        assert "СТР-11 О /2/" not in base_groups

        # group_structure should indicate subgroups
        assert group_structure["СТР-11 О"]["has_subgroups"] is True

    def test_merge_vertically_helper(self):
        """Test the _merge_vertically helper method."""
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)

        result = generator._merge_vertically("Top\nLine2", "Bottom\nLine2")
        lines = result.split("\n")
        assert len(lines) == 5  # Top(2) + separator(1) + Bottom(2)
        assert "Top" == lines[0]
        assert "Line2" == lines[1]
        assert "-" == lines[2]
        assert "Bottom" == lines[3]
        assert "Line2" == lines[4]

    def test_merge_vertically_empty_top(self):
        """Test vertical merge with empty top side."""
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)

        result = generator._merge_vertically("", "Bottom\nLine2\nLine3")
        lines = result.split("\n")
        assert len(lines) == 4  # separator(1) + Bottom(3)
        assert "-" == lines[0]
        assert "Bottom" == lines[1]

    def test_merge_vertically_empty_bottom(self):
        """Test vertical merge with empty bottom side."""
        config = GeneratorConfig(language="kaz", year=1, week_type="both")
        generator = ScheduleExcelGenerator(config)

        result = generator._merge_vertically("Top\nLine2", "")
        lines = result.split("\n")
        assert len(lines) == 3  # Top(2) + separator(1)
        assert "Top" == lines[0]
        assert "Line2" == lines[1]
        assert "-" == lines[2]
