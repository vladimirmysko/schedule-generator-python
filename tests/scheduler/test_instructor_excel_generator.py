"""Tests for instructor Excel schedule generator."""

import json
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from form1_parser.scheduler.instructor_excel_generator import (
    DAYS_ORDER,
    FILL_BOTH_WEEKS,
    FILL_EVEN_WEEK,
    FILL_ODD_WEEK,
    InstructorGeneratorConfig,
    InstructorScheduleExcelGenerator,
    generate_instructor_schedule_excel,
)


# Sample test data
SAMPLE_ASSIGNMENTS = [
    {
        "stream_id": "test_stream_1",
        "subject": "Test Subject 1",
        "instructor": "а.о.Instructor One",
        "groups": ["АРХ-11 О", "АРХ-13 О"],
        "student_count": 50,
        "day": "monday",
        "slot": 1,
        "time": "09:00-09:50",
        "room": "A-101",
        "room_address": "Test Address 1",
        "week_type": "both",
        "stream_type": "lecture",
    },
    {
        "stream_id": "test_stream_2",
        "subject": "Test Subject 2",
        "instructor": "а.о.Instructor One",
        "groups": ["АРХ-11 О"],
        "student_count": 25,
        "day": "monday",
        "slot": 2,
        "time": "10:00-10:50",
        "room": "B-202",
        "room_address": "Test Address 2",
        "week_type": "odd",
        "stream_type": "practical",
    },
    {
        "stream_id": "test_stream_3",
        "subject": "Test Subject 3",
        "instructor": "қ.проф.Instructor Two",
        "groups": ["ЮР-12 О", "ЮР-14 О"],
        "student_count": 30,
        "day": "wednesday",
        "slot": 6,
        "time": "14:00-14:50",
        "room": "C-303",
        "room_address": "Test Address 3",
        "week_type": "even",
        "stream_type": "lab",
    },
    {
        "stream_id": "test_stream_4",
        "subject": "Test Subject 4",
        "instructor": "а.о.Instructor One",
        "groups": ["АРХ-21 О"],
        "student_count": 20,
        "day": "tuesday",
        "slot": 3,
        "time": "11:00-11:50",
        "room": "D-404",
        "room_address": "Test Address 4",
        "week_type": "both",
        "stream_type": "lecture",
    },
    {
        "stream_id": "test_stream_5",
        "subject": "Test Subject 5",
        "instructor": "а.о.Instructor One",
        "groups": ["АРХ-11 О"],
        "student_count": 25,
        "day": "monday",
        "slot": 2,
        "time": "10:00-10:50",
        "room": "E-505",
        "room_address": "Test Address 5",
        "week_type": "even",
        "stream_type": "practical",
    },
]

SAMPLE_SCHEDULE_DATA = {
    "generation_date": "2025-01-20T10:00:00",
    "stage": 1,
    "total_assigned": 5,
    "total_unscheduled": 0,
    "assignments": SAMPLE_ASSIGNMENTS,
    "unscheduled_stream_ids": [],
    "statistics": {},
}


class TestGroupByInstructor:
    """Tests for grouping assignments by instructor."""

    def test_group_by_instructor(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        by_instructor = generator.group_by_instructor(SAMPLE_SCHEDULE_DATA)

        # Names are normalized (prefixes removed, spacing normalized)
        assert "Instructor One" in by_instructor
        assert "Instructor Two" in by_instructor
        assert len(by_instructor["Instructor One"]) == 4
        assert len(by_instructor["Instructor Two"]) == 1

    def test_group_by_instructor_empty_data(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        by_instructor = generator.group_by_instructor({"assignments": []})

        assert len(by_instructor) == 0

    def test_group_by_instructor_merges_inconsistent_spacing(self):
        """Test that instructors with inconsistent spacing are merged."""
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)

        # Same instructor with different spacing after prefix
        data = {
            "assignments": [
                {"instructor": "а.о. Шалаев Б.Б.", "day": "monday", "slot": 1},
                {"instructor": "а.о.Шалаев Б.Б.", "day": "friday", "slot": 2},
            ]
        }
        by_instructor = generator.group_by_instructor(data)

        # Should be merged into one instructor
        assert len(by_instructor) == 1
        assert "Шалаев Б.Б." in by_instructor
        assert len(by_instructor["Шалаев Б.Б."]) == 2


class TestBuildInstructorGrid:
    """Tests for building instructor schedule grid."""

    def test_grid_structure(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignments = [a for a in SAMPLE_ASSIGNMENTS if "Instructor One" in a["instructor"]]
        grid = generator.build_instructor_grid(assignments)

        # Check all days are present
        for day in DAYS_ORDER:
            assert day in grid

        # Check all slots (1-13) are present
        for slot in range(1, 14):
            assert slot in grid["monday"]

    def test_grid_assignment_placement(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignments = [a for a in SAMPLE_ASSIGNMENTS if "Instructor One" in a["instructor"]]
        grid = generator.build_instructor_grid(assignments)

        # Check monday slot 1 has assignment
        assert len(grid["monday"][1]) == 1
        assert grid["monday"][1][0]["subject"] == "Test Subject 1"

        # Check tuesday slot 3 has assignment
        assert len(grid["tuesday"][3]) == 1
        assert grid["tuesday"][3][0]["subject"] == "Test Subject 4"

    def test_grid_multiple_assignments_same_slot(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignments = [a for a in SAMPLE_ASSIGNMENTS if "Instructor One" in a["instructor"]]
        grid = generator.build_instructor_grid(assignments)

        # Monday slot 2 has two assignments (odd and even weeks)
        assert len(grid["monday"][2]) == 2


class TestFormatCellContent:
    """Tests for cell content formatting."""

    def test_format_lecture(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignment = {
            "subject": "Test Subject",
            "stream_type": "lecture",
            "groups": ["АРХ-11 О", "АРХ-13 О"],
            "room": "A-101",
            "room_address": "Test Address",
        }
        content = generator.format_cell_content(assignment)

        # Lectures should be uppercase
        assert "TEST SUBJECT" in content
        assert "[Дәріс]" in content  # Kazakh for "Lecture"
        assert "АРХ-11 О, АРХ-13 О" in content
        assert "A-101, Test Address" in content

    def test_format_practical(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignment = {
            "subject": "Test Subject",
            "stream_type": "practical",
            "groups": ["АРХ-11 О"],
            "room": "B-202",
            "room_address": "Test Address",
        }
        content = generator.format_cell_content(assignment)

        # Practicals should not be uppercase
        assert "Test Subject" in content
        assert "[Практика]" in content

    def test_format_lab(self):
        config = InstructorGeneratorConfig(language="rus")
        generator = InstructorScheduleExcelGenerator(config)
        assignment = {
            "subject": "Test Subject",
            "stream_type": "lab",
            "groups": ["ЮР-12 О"],
            "room": "C-303",
            "room_address": "Test Address",
        }
        content = generator.format_cell_content(assignment)

        assert "Test Subject" in content
        assert "[Лаборатория]" in content  # Russian for "Lab"


class TestGetCellFill:
    """Tests for cell fill color determination."""

    def test_both_weeks(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignments = [{"week_type": "both"}]
        fill = generator.get_cell_fill(assignments)

        assert fill == FILL_BOTH_WEEKS

    def test_odd_week_only(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignments = [{"week_type": "odd"}]
        fill = generator.get_cell_fill(assignments)

        assert fill == FILL_ODD_WEEK

    def test_even_week_only(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignments = [{"week_type": "even"}]
        fill = generator.get_cell_fill(assignments)

        assert fill == FILL_EVEN_WEEK

    def test_odd_and_even_together(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        assignments = [{"week_type": "odd"}, {"week_type": "even"}]
        fill = generator.get_cell_fill(assignments)

        # Both odd and even specific = white (both weeks)
        assert fill == FILL_BOTH_WEEKS

    def test_empty_assignments(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        fill = generator.get_cell_fill([])

        assert fill == FILL_BOTH_WEEKS


class TestSanitizeSheetName:
    """Tests for sheet name sanitization."""

    def test_removes_instructor_prefix(self):
        name = InstructorScheduleExcelGenerator.sanitize_sheet_name("а.о.Instructor Name")
        assert "а.о." not in name
        assert "Instructor Name" in name

    def test_removes_professor_prefix(self):
        name = InstructorScheduleExcelGenerator.sanitize_sheet_name("қ.проф.Professor Name")
        assert "қ.проф." not in name
        assert "Professor Name" in name

    def test_removes_invalid_characters(self):
        name = InstructorScheduleExcelGenerator.sanitize_sheet_name("Name/With:Invalid*Chars")
        assert "/" not in name
        assert ":" not in name
        assert "*" not in name

    def test_truncates_long_names(self):
        long_name = "A" * 50
        name = InstructorScheduleExcelGenerator.sanitize_sheet_name(long_name)
        assert len(name) <= 31


class TestCleanInstructorName:
    """Tests for instructor name cleaning."""

    def test_removes_ao_prefix(self):
        name = InstructorScheduleExcelGenerator.clean_instructor_name("а.о.Instructor Name")
        assert name == "Instructor Name"

    def test_removes_professor_prefix(self):
        name = InstructorScheduleExcelGenerator.clean_instructor_name("қ.проф.Professor Name")
        assert name == "Professor Name"

    def test_strips_whitespace(self):
        name = InstructorScheduleExcelGenerator.clean_instructor_name("  Name  ")
        assert name == "Name"


class TestCreateWorkbook:
    """Tests for workbook creation."""

    def test_creates_workbook_with_sheets(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        wb = generator.create_workbook(SAMPLE_SCHEDULE_DATA)

        # Should have one sheet per instructor
        assert len(wb.worksheets) == 2

    def test_sheets_named_by_instructor(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        wb = generator.create_workbook(SAMPLE_SCHEDULE_DATA)

        sheet_names = [ws.title for ws in wb.worksheets]
        # Names should be sanitized (prefixes removed)
        assert "Instructor One" in sheet_names
        assert "Instructor Two" in sheet_names

    def test_empty_data_creates_empty_sheet(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        wb = generator.create_workbook({"assignments": []})

        assert len(wb.worksheets) == 1
        assert "Empty" in wb.worksheets[0].title


class TestSheetLayout:
    """Tests for sheet layout and formatting."""

    def test_sheet_has_university_name(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        wb = generator.create_workbook(SAMPLE_SCHEDULE_DATA)

        ws = wb.worksheets[0]
        assert ws["A2"].value is not None
        assert "университет" in ws["A2"].value.lower()

    def test_sheet_has_schedule_title(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        wb = generator.create_workbook(SAMPLE_SCHEDULE_DATA)

        ws = wb.worksheets[0]
        assert ws["A4"].value is not None
        assert "КЕСТЕ" in ws["A4"].value.upper()

    def test_sheet_has_instructor_name(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        wb = generator.create_workbook(SAMPLE_SCHEDULE_DATA)

        ws = wb.worksheets[0]
        # Row 5 should have instructor name
        assert ws["A5"].value is not None

    def test_sheet_has_slot_headers(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        wb = generator.create_workbook(SAMPLE_SCHEDULE_DATA)

        ws = wb.worksheets[0]
        # Row 9 should have headers
        assert ws["A9"].value is not None  # Day header
        # Check slot columns have time ranges
        assert "09:00" in str(ws["B9"].value)  # Slot 1
        assert "10:00" in str(ws["C9"].value)  # Slot 2

    def test_sheet_has_day_names(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)
        wb = generator.create_workbook(SAMPLE_SCHEDULE_DATA)

        ws = wb.worksheets[0]
        strings = generator.strings

        # Check day names in column A (rows 10-14)
        day_values = [ws[f"A{row}"].value for row in range(10, 15)]
        for day in DAYS_ORDER:
            assert strings["days"][day] in day_values


class TestLocalization:
    """Tests for language localization."""

    def test_kazakh_strings(self):
        config = InstructorGeneratorConfig(language="kaz")
        generator = InstructorScheduleExcelGenerator(config)

        assert generator.strings["schedule_title"] == "ОҚЫТУШЫ КЕСТЕСІ"
        assert generator.strings["legend_odd"] == "Тақ апта"

    def test_russian_strings(self):
        config = InstructorGeneratorConfig(language="rus")
        generator = InstructorScheduleExcelGenerator(config)

        assert generator.strings["schedule_title"] == "РАСПИСАНИЕ ПРЕПОДАВАТЕЛЯ"
        assert generator.strings["legend_odd"] == "Нечетная неделя"


class TestGenerateInstructorScheduleExcel:
    """Tests for the main generation function."""

    def test_generate_single_language(self):
        with TemporaryDirectory() as tmpdir:
            # Write sample JSON
            input_path = Path(tmpdir) / "schedule.json"
            with open(input_path, "w", encoding="utf-8") as f:
                json.dump(SAMPLE_SCHEDULE_DATA, f)

            output_dir = Path(tmpdir) / "output"

            # Generate single language
            files = generate_instructor_schedule_excel(
                input_path=input_path,
                output_dir=output_dir,
                language="kaz",
            )

            assert len(files) == 1
            assert files[0].exists()
            assert "instructor_schedules_kaz.xlsx" in str(files[0])

            # Verify workbook can be opened
            wb = load_workbook(files[0])
            assert len(wb.worksheets) == 2

    def test_generate_both_languages(self):
        with TemporaryDirectory() as tmpdir:
            # Write sample JSON
            input_path = Path(tmpdir) / "schedule.json"
            with open(input_path, "w", encoding="utf-8") as f:
                json.dump(SAMPLE_SCHEDULE_DATA, f)

            output_dir = Path(tmpdir) / "output"

            # Generate both languages (no filter)
            files = generate_instructor_schedule_excel(
                input_path=input_path,
                output_dir=output_dir,
            )

            assert len(files) == 2
            filenames = [f.name for f in files]
            assert "instructor_schedules_kaz.xlsx" in filenames
            assert "instructor_schedules_rus.xlsx" in filenames

    def test_output_directory_created(self):
        with TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "schedule.json"
            with open(input_path, "w", encoding="utf-8") as f:
                json.dump(SAMPLE_SCHEDULE_DATA, f)

            # Use nested directory that doesn't exist
            output_dir = Path(tmpdir) / "nested" / "output" / "dir"

            files = generate_instructor_schedule_excel(
                input_path=input_path,
                output_dir=output_dir,
                language="kaz",
            )

            assert output_dir.exists()
            assert len(files) == 1


class TestScheduleContent:
    """Tests for actual schedule content in generated files."""

    def test_schedule_contains_assignments(self):
        with TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "schedule.json"
            with open(input_path, "w", encoding="utf-8") as f:
                json.dump(SAMPLE_SCHEDULE_DATA, f)

            output_dir = Path(tmpdir) / "output"

            files = generate_instructor_schedule_excel(
                input_path=input_path,
                output_dir=output_dir,
                language="kaz",
            )

            wb = load_workbook(files[0])
            ws = wb["Instructor One"]

            # Check that the schedule has content
            # Row 10 is Monday, column B is slot 1
            assert ws["B10"].value is not None
            assert "TEST SUBJECT 1" in ws["B10"].value.upper()

    def test_week_type_colors_applied(self):
        with TemporaryDirectory() as tmpdir:
            input_path = Path(tmpdir) / "schedule.json"
            with open(input_path, "w", encoding="utf-8") as f:
                json.dump(SAMPLE_SCHEDULE_DATA, f)

            output_dir = Path(tmpdir) / "output"

            files = generate_instructor_schedule_excel(
                input_path=input_path,
                output_dir=output_dir,
                language="kaz",
            )

            wb = load_workbook(files[0])
            ws = wb["Instructor Two"]

            # Wednesday (row 12), slot 6 (column G) should have even week assignment
            cell = ws["G12"]
            assert cell.value is not None
            # Cell should have even week fill color (orange)
            assert cell.fill.start_color.rgb == FILL_EVEN_WEEK.start_color.rgb
